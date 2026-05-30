"""
file_ops.py — NEXUS File Operations
Handles reading, summarising, and answering questions about files.
Supports .txt, .pdf, and .xlsx formats. All file access is gated
through permission_guard.is_allowed() for security.
"""

import os
import shutil
import pdfplumber
import openpyxl

from model_loader import generate
from permission_guard import is_allowed, load_permissions
from utils import chunk_text, clean_text, format_size
import ui


# ---------------------------------------------------------------------------
# Known error prefixes (used to detect reader failures vs real content)
# ---------------------------------------------------------------------------
_ERROR_PREFIXES = (
    "Permission denied",
    "File not found",
    "Unsupported file type",
    "Error reading",
    "Error:",
)


# ---------------------------------------------------------------------------
# Path resolution — find files inside allowed folders
# ---------------------------------------------------------------------------
def resolve_filepath(filepath):
    """
    Resolve a filepath to an absolute path. If the path is relative
    (e.g. just a filename like 'test.txt'), search for it inside all
    allowed folders and return the first match found.

    Returns the resolved absolute path, or the original path if no
    match is found (let the caller handle the error).
    """
    # Already absolute — return as-is
    if os.path.isabs(filepath):
        return filepath

    # Check if it exists relative to current working directory first
    cwd_path = os.path.abspath(filepath)
    if os.path.exists(cwd_path) and is_allowed(cwd_path):
        return cwd_path

    # Search through all allowed folders for this filename
    allowed_paths = load_permissions()
    for allowed_dir in allowed_paths:
        candidate = os.path.join(os.path.abspath(allowed_dir), filepath)
        if os.path.exists(candidate):
            return candidate

    # Nothing found — return the original as-is so callers get a clear error
    return filepath


# ---------------------------------------------------------------------------
# Individual file readers
# ---------------------------------------------------------------------------
def read_txt(filepath):
    """
    Read a plain text file and return its content as a string.
    Checks permissions before access.
    """
    if not is_allowed(filepath):
        return "Permission denied for this file."

    if not os.path.isfile(filepath):
        return "File not found."

    try:
        with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    except Exception as e:
        return f"Error reading text file: {e}"


def read_pdf(filepath):
    """
    Extract text from a PDF file using pdfplumber.
    Returns concatenated text from all pages with page separators.
    """
    if not is_allowed(filepath):
        return "Permission denied for this file."

    if not os.path.isfile(filepath):
        return "File not found."

    try:
        pages_text = []
        with pdfplumber.open(filepath) as pdf:
            for i, page in enumerate(pdf.pages, 1):
                text = page.extract_text()
                if text:
                    pages_text.append(f"--- Page {i} ---\n{text}")
                else:
                    pages_text.append(f"--- Page {i} --- [Image-only page, no text extracted]")

        if not pages_text:
            return "Error: Could not extract any text from this PDF."

        return "\n\n".join(pages_text)
    except Exception as e:
        return f"Error reading PDF file: {e}"


def read_excel(filepath):
    """
    Convert an Excel file to a plain text table representation.
    Reads all sheets and joins cell values with ' | ' separators.
    """
    if not is_allowed(filepath):
        return "Permission denied for this file."

    if not os.path.isfile(filepath):
        return "File not found."

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        all_text = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_text.append(f"=== Sheet: {sheet_name} ===")

            for row in ws.iter_rows(values_only=True):
                # Skip entirely empty rows
                if all(cell is None for cell in row):
                    continue
                # Join cell values, replacing None with empty string
                row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                all_text.append(row_text)

            all_text.append("")  # blank line between sheets

        wb.close()
        return "\n".join(all_text)
    except Exception as e:
        return f"Error reading Excel file: {e}"


# ---------------------------------------------------------------------------
# Dispatcher — picks the right reader based on file extension
# ---------------------------------------------------------------------------
def read_file(filepath):
    """
    Automatically select the correct reader based on file extension.
    Supports: .txt, .pdf, .xlsx, .xls
    """
    ext = os.path.splitext(filepath)[1].lower()

    if ext == ".txt":
        return read_txt(filepath)
    elif ext == ".pdf":
        return read_pdf(filepath)
    elif ext in (".xlsx", ".xls"):
        return read_excel(filepath)
    else:
        return "Unsupported file type. Supported: .txt, .pdf, .xlsx"


# ---------------------------------------------------------------------------
# Summarise a single file
# ---------------------------------------------------------------------------
def summarise_file(filepath):
    """
    Read a file and produce an AI-generated summary.
    Handles long files by chunking and summarising each chunk,
    then combining into a final summary.
    """
    # Resolve relative paths — search allowed folders for the file
    filepath = resolve_filepath(filepath)

    raw_text = read_file(filepath)

    # If the reader returned an error, pass it through
    if raw_text.startswith(_ERROR_PREFIXES):
        return raw_text

    # Clean and chunk
    cleaned = clean_text(raw_text)

    if not cleaned:
        return "The file appears to be empty or contains no readable text."

    chunks = chunk_text(cleaned)

    if len(chunks) == 0:
        return "The file appears to be empty or contains no readable text."

    if len(chunks) == 1:
        # Single chunk — summarise directly
        prompt = "Summarise the following text in 3-5 sentences. Be brief:\n\n" + chunks[0]
        return generate(prompt, max_tokens=250)

    # Multiple chunks — summarise each, then combine
    ui.print_status(f"File is large ({len(chunks)} chunks). Summarising each chunk...")
    chunk_summaries = []
    for i, chunk in enumerate(chunks, 1):
        ui.print_status(f"Processing chunk {i}/{len(chunks)}...")
        prompt = "Summarise the following text in 3-5 sentences. Be brief:\n\n" + chunk
        summary = generate(prompt, max_tokens=250)
        chunk_summaries.append(summary)

    # Combine all chunk summaries into one final summary
    combined = "\n\n".join(
        f"[Chunk {i}]: {s}" for i, s in enumerate(chunk_summaries, 1)
    )
    final_prompt = "Combine these summaries into one coherent summary in 3-5 sentences:\n\n" + combined
    return generate(final_prompt, max_tokens=250)


# ---------------------------------------------------------------------------
# Summarise all supported files in a folder
# ---------------------------------------------------------------------------
def summarise_folder(folderpath):
    """
    Summarise all supported files (.txt, .pdf, .xlsx) in a folder.
    Produces individual summaries and then a master summary.
    Saves output to summary_output.txt in the same folder.
    """
    # Resolve relative paths — search allowed folders
    if not os.path.isabs(folderpath):
        # Try to find it as a subfolder of an allowed path
        allowed_paths = load_permissions()
        for allowed_dir in allowed_paths:
            candidate = os.path.join(os.path.abspath(allowed_dir), folderpath)
            if os.path.isdir(candidate):
                folderpath = candidate
                break
        else:
            folderpath = os.path.abspath(folderpath)

    if not is_allowed(folderpath):
        return "Permission denied for this folder."

    if not os.path.isdir(folderpath):
        return "Folder not found."

    # Find supported files
    supported_exts = (".txt", ".pdf", ".xlsx", ".xls")
    files = [
        f for f in os.listdir(folderpath)
        if os.path.isfile(os.path.join(folderpath, f))
        and f.lower().endswith(supported_exts)
        and f != "summary_output.txt"  # don't summarise our own output
    ]

    if not files:
        return "No supported files found in this folder. Supported: .txt, .pdf, .xlsx"

    # Summarise each file
    summaries = {}
    for filename in files:
        filepath = os.path.join(folderpath, filename)
        ui.print_status(f"Processing {filename}...")
        summaries[filename] = summarise_file(filepath)

    # Build combined text
    combined_parts = []
    for filename, summary in summaries.items():
        combined_parts.append(f"=== {filename} ===\n{summary}")

    combined = "\n\n".join(combined_parts)

    # Generate master summary
    ui.print_status("Generating master summary...")
    master_prompt = "Create one master summary combining all these documents:\n\n" + combined
    master_summary = generate(master_prompt)

    # Build the full output
    output_text = "NEXUS Folder Summary\n"
    output_text += "=" * 50 + "\n\n"
    for filename, summary in summaries.items():
        output_text += f"--- {filename} ---\n{summary}\n\n"
    output_text += "=" * 50 + "\n"
    output_text += "MASTER SUMMARY\n"
    output_text += "=" * 50 + "\n"
    output_text += master_summary + "\n"

    # Save to file
    output_path = os.path.join(folderpath, "summary_output.txt")
    try:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(output_text)
        return f"{master_summary}\n\n[Full summary saved to: {output_path}]"
    except IOError as e:
        return f"{master_summary}\n\n[Could not save summary file: {e}]"

# ---------------------------------------------------------------------------
# File management — rename, copy, move, delete, list (Phase 18)
# ---------------------------------------------------------------------------
def rename_file(old_path, new_name):
    """
    Rename a file. old_path is the full or relative path to the file,
    new_name is just the new filename (not a full path).
    Checks permissions, asks for confirmation before executing.
    """
    old_path = resolve_filepath(old_path)

    if not is_allowed(old_path):
        return "Permission denied for this file."
    if not os.path.isfile(old_path):
        return f"File not found: {old_path}"

    directory = os.path.dirname(old_path)
    new_path = os.path.join(directory, new_name)

    if os.path.exists(new_path):
        return f"Cannot rename: a file named '{new_name}' already exists in that folder."

    old_name = os.path.basename(old_path)
    print(f"\n  Rename: {old_name}  -->  {new_name}")
    print(f"  In folder: {directory}")
    confirm = input("  Proceed? (y/n): ").strip().lower()

    if confirm != "y":
        return "Rename cancelled."

    try:
        os.rename(old_path, new_path)
        return f"File renamed successfully: {old_name} --> {new_name}"
    except Exception as e:
        return f"Error renaming file: {e}"


def copy_file(source, destination):
    """
    Copy a file from source to destination.
    Checks permissions for both source and destination.
    Asks for confirmation before executing.
    """
    source = resolve_filepath(source)

    if not is_allowed(source):
        return "Permission denied for the source file."
    if not os.path.isfile(source):
        return f"Source file not found: {source}"

    # If destination is just a folder, keep the original filename
    if os.path.isdir(destination):
        destination = os.path.join(destination, os.path.basename(source))

    dest_dir = os.path.dirname(os.path.abspath(destination))
    if not is_allowed(dest_dir):
        return "Permission denied for the destination folder."

    print(f"\n  Copy: {os.path.basename(source)}")
    print(f"  From: {os.path.dirname(source)}")
    print(f"  To:   {dest_dir}")
    confirm = input("  Proceed? (y/n): ").strip().lower()

    if confirm != "y":
        return "Copy cancelled."

    try:
        shutil.copy2(source, destination)
        return f"File copied successfully to: {destination}"
    except Exception as e:
        return f"Error copying file: {e}"


def move_file(source, destination):
    """
    Move a file from source to destination.
    Checks permissions for both source and destination.
    Asks for confirmation before executing.
    """
    source = resolve_filepath(source)

    if not is_allowed(source):
        return "Permission denied for the source file."
    if not os.path.isfile(source):
        return f"Source file not found: {source}"

    # If destination is just a folder, keep the original filename
    if os.path.isdir(destination):
        destination = os.path.join(destination, os.path.basename(source))

    dest_dir = os.path.dirname(os.path.abspath(destination))
    if not is_allowed(dest_dir):
        return "Permission denied for the destination folder."

    print(f"\n  Move: {os.path.basename(source)}")
    print(f"  From: {os.path.dirname(source)}")
    print(f"  To:   {dest_dir}")
    confirm = input("  Proceed? (y/n): ").strip().lower()

    if confirm != "y":
        return "Move cancelled."

    try:
        shutil.move(source, destination)
        return f"File moved successfully to: {destination}"
    except Exception as e:
        return f"Error moving file: {e}"


def delete_file(filepath):
    """
    Delete a file. Asks for extra confirmation since this is destructive.
    Checks permissions before executing.
    """
    filepath = resolve_filepath(filepath)

    if not is_allowed(filepath):
        return "Permission denied for this file."
    if not os.path.isfile(filepath):
        return f"File not found: {filepath}"

    filename = os.path.basename(filepath)
    size_bytes = os.path.getsize(filepath)
    size_str = format_size(size_bytes)

    print(f"\n  [WARNING] DELETE FILE: {filename} ({size_str})")
    print(f"  Location: {os.path.dirname(filepath)}")
    print(f"  This action cannot be undone!")
    confirm = input("  Are you sure? (y/n): ").strip().lower()

    if confirm != "y":
        return "Delete cancelled."

    try:
        os.remove(filepath)
        return f"File deleted: {filename}"
    except Exception as e:
        return f"Error deleting file: {e}"


def list_files(folderpath):
    """
    List all files in a folder with their sizes.
    Checks permissions before listing.
    """
    # Resolve relative paths
    if not os.path.isabs(folderpath):
        allowed_paths = load_permissions()
        for allowed_dir in allowed_paths:
            candidate = os.path.join(os.path.abspath(allowed_dir), folderpath)
            if os.path.isdir(candidate):
                folderpath = candidate
                break
        else:
            folderpath = os.path.abspath(folderpath)

    if not is_allowed(folderpath):
        return "Permission denied for this folder."
    if not os.path.isdir(folderpath):
        return f"Folder not found: {folderpath}"

    entries = []
    try:
        for item in sorted(os.listdir(folderpath)):
            full_path = os.path.join(folderpath, item)
            if os.path.isfile(full_path):
                size = format_size(os.path.getsize(full_path))
                entries.append(f"  [FILE] {item:<40s}  {size}")
            elif os.path.isdir(full_path):
                entries.append(f"  [DIR]  {item}")
    except PermissionError:
        return "Permission denied by the operating system for this folder."

    if not entries:
        return f"Folder is empty: {folderpath}"

    header = f"Contents of: {folderpath}\n{'=' * 60}\n"
    return header + "\n".join(entries)


